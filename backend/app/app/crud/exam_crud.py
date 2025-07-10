from collections import defaultdict
from ctypes import alignment
import os
from fastapi.responses import FileResponse
from sqlalchemy import desc
from sqlalchemy.orm import Session
from datetime import datetime
from app.models import Exam,Marks,ExamResult,QuestionPaper,AllocatedMember,Allocation,User,Subject
from app.schemas import ExamSchema,MarksSchema,ExamResultSchema,QuestionPaperSchema
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font as ExcelFont, Alignment
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


#-----------------------------------CRUD for exams---------------------------------------------------------#

def create_bulk_exams(db: Session, exam_data: list[ExamSchema]):
    now = datetime.utcnow()
    exam_objs = []

    for data in exam_data:
        exam = Exam(
            allocation_id=data.allocation_id,
            exam_name=data.exam_name,
            exam_code=data.exam_code,
            total_marks=data.total_marks,
            start_date=data.start_date,
            end_date=data.end_date,
            status=data.status,
            created_by=data.created_by,
            updated_by=data.updated_by,
            created_at=now,
            updated_at=now
        )
        exam_objs.append(exam)

    db.add_all(exam_objs)
    db.commit()
    return exam_objs

def get_exams_by_allocation_ids(db: Session, allocation_ids: list[int]):
    return db.query(Exam).filter(Exam.allocation_id.in_(allocation_ids), Exam.status == 1).all()

def get_exam_by_id(db: Session, id: int):
    return db.query(Exam).filter(Exam.id == id).first()

def update_exam(db: Session, id: int, data: ExamSchema):
    db_obj = db.query(Exam).filter(Exam.id == id).first()
    if not db_obj:
        return None
    for field, value in data.dict(exclude_unset=True).items():
        setattr(db_obj, field, value)
    db_obj.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(db_obj)
    return db_obj

def soft_delete_exam(db: Session, id: int, updated_by: int):
    db_obj = db.query(Exam).filter(Exam.id == id).first()
    if db_obj:
        db_obj.status = -1
        db_obj.updated_by = updated_by
        db_obj.updated_at = datetime.utcnow()
        db.commit()
        return True
    return False

#------------------------------------------CRUD for marks-----------------------------------------------------#

def create_bulk_marks(db: Session, marks_data: list[MarksSchema]):
    now = datetime.utcnow()
    marks_objs = []

    for data in marks_data:
        mark = Marks(
            student_id=data.student_id,
            exam_id=data.exam_id,
            subject_id=data.subject_id,
            obtained_marks=data.obtained_marks,
            status=data.status,
            created_by=data.created_by,
            updated_by=data.updated_by,
            created_at=now,
            updated_at=now
        )
        marks_objs.append(mark)

    db.add_all(marks_objs)
    db.commit()
    return marks_objs

def get_all_marks(db: Session):
    return db.query(Marks).filter(Marks.status == 1).all()

def get_mark_by_id(db: Session, id: int):
    return db.query(Marks).filter(Marks.id == id).first()

def update_mark(db: Session, id: int, data: MarksSchema):
    db_obj = db.query(Marks).filter(Marks.id == id).first()
    if not db_obj:
        return None
    for field, value in data.dict(exclude_unset=True).items():
        setattr(db_obj, field, value)
    db_obj.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(db_obj)
    return db_obj

def soft_delete_mark(db: Session, id: int, updated_by: int):
    db_obj = db.query(Marks).filter(Marks.id == id).first()
    if db_obj:
        db_obj.status = -1
        db_obj.updated_by = updated_by
        db_obj.updated_at = datetime.utcnow()
        db.commit()
        return True
    return False

#--------------------------------------CRUD for exam_result-------------------------------------------#

def create_bulk_exam_results(db: Session, results_data: list[ExamResultSchema]):
    now = datetime.utcnow()
    result_objs = []

    for data in results_data:
        result = ExamResult(
            student_id=data.student_id,
            exam_id=data.exam_id,
            allocation_id=data.allocation_id,
            total_obtained=data.total_obtained,
            status=data.status,
            created_by=data.created_by,
            updated_by=data.updated_by,
            created_at=now,
            updated_at=now
        )
        result_objs.append(result)

    db.add_all(result_objs)
    db.commit()
    return result_objs

def get_all_exam_results(db: Session):
    return db.query(ExamResult).all()

def get_exam_result_by_student(db: Session, id: int):
    return db.query(ExamResult).filter(ExamResult.student_id == id).first()

def get_results_by_class_teacher(db: Session, teacher_id: int):
    allocation_ids = (
        db.query(AllocatedMember.allocation_id)
        .filter(
            AllocatedMember.user_id == teacher_id,
            AllocatedMember.role == "class_teacher"
        )
        .all()
    )

    allocation_ids = [row[0] for row in allocation_ids]
    if not allocation_ids:
        return []

    results = (
        db.query(ExamResult)
        .join(User, ExamResult.student_id == User.id)
        .join(Exam, ExamResult.exam_id == Exam.id)
        .filter(ExamResult.allocation_id.in_(allocation_ids))
        .order_by(ExamResult.student_id)
        .all()
    )

    return results


def update_exam_result(db: Session, id: int, data: ExamResultSchema):
    db_obj = db.query(ExamResult).filter(ExamResult.id == id).first()
    if not db_obj:
        return None
    for field, value in data.dict(exclude_unset=True).items():
        setattr(db_obj, field, value)
    db_obj.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(db_obj)
    return db_obj

def delete_exam_result(db: Session, id: int):
    db_obj = db.query(ExamResult).filter(ExamResult.id == id).first()
    if db_obj:
        db.delete(db_obj)
        db.commit()
        return True
    return False

def get_student_marks_with_rank(db: Session, student_id: int, exam_id: int):
    student_result = db.query(ExamResult).filter(
        ExamResult.student_id == student_id,
        ExamResult.exam_id == exam_id
    ).first()

    if not student_result:
        return {"status": 0, "detail": "Student result not found in this exam"}

    all_totals = (
        db.query(ExamResult.student_id, ExamResult.total_obtained)
        .filter(
            ExamResult.exam_id == exam_id,
            ExamResult.allocation_id == student_result.allocation_id
        )
        .order_by(ExamResult.total_obtained.desc())
        .all()
    )

    rank = None
    for idx, (sid, total) in enumerate(all_totals, start=1):
        if sid == student_id:
            rank = idx
            break

    if rank is None:
        return {"status": 0, "detail": "Student rank not found in this exam"}

    # Step 3: Get subject-wise marks
    subject_marks = (
        db.query(
            Subject.subject_name.label("subject"),
            Marks.obtained_marks,
            Marks.obtained_marks
        )
        .join(Marks, Marks.subject_id == Subject.id)
        .filter(
            Marks.student_id == student_id,
            Marks.exam_id == exam_id
        )
        .all()
    )

    return {
        "status": 1,
        "data": {
            "student_id": student_id,
            "exam_id": exam_id,
            "total_obtained": student_result.total_obtained,
            "rank": rank,
            "subject_marks": [
                {
                    "subject": sub,
                    "total_marks": total,
                    "obtained_marks": obtained
                }
                for sub, total, obtained in subject_marks
            ]
        }
    }

def generate_student_marksheet(student_id: int, db: Session):
    student = db.query(User).filter(User.id == student_id).first()
    if not student:
        return {"status":0, "detail":"Student not found"}

    exam_results = db.query(ExamResult).filter(ExamResult.student_id == student_id).all()
    if not exam_results:
        return {"status":0, "detail":"No exam results found for this student"}

    final_data = []

    for er in exam_results:
        exam = er.exam
        if not exam:
            continue

        subject_marks = (
            db.query(
                Subject.subject_name,
                exam.total_marks,
                Marks.obtained_marks
            )
            .join(Marks, Marks.subject_id == Subject.id)
            .filter(Marks.student_id == student_id, Marks.exam_id == exam.id)
            .all()
        )

        all_totals = (
            db.query(ExamResult.student_id, ExamResult.total_obtained)
            .filter(ExamResult.exam_id == exam.id)
            .order_by(desc(ExamResult.total_obtained))
            .all()
        )

        rank = None
        for idx, (sid, _) in enumerate(all_totals, start=1):
            if sid == student_id:
                rank = idx
                break

        final_data.append({
            "exam_name": exam.exam_name,
            "total_marks": exam.total_marks,
            "obtained_marks": er.total_obtained,
            "rank": rank,
            "subjects": [
                {
                    "subject": s,
                    "total_marks": tm,
                    "obtained_marks": om
                } for s, tm, om in subject_marks
            ]
        })

    # Generate PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt=f"Mark Sheet for {student.first_name}", ln=True, align="C")

    for exam in final_data:
        pdf.ln(10)
        pdf.set_font("Arial", "B", size=12)
        pdf.cell(200, 10, txt=f"{exam['exam_name'].upper()} RESULT", ln=True)
        pdf.set_font("Arial", size=12)
        pdf.cell(100, 10, txt=f"Total Marks: {exam['total_marks']}", ln=True)
        pdf.cell(100, 10, txt=f"Obtained Marks: {exam['obtained_marks']}", ln=True)
        pdf.cell(100, 10, txt=f"Rank: {exam['rank']}", ln=True)
        pdf.ln(5)
        for sub in exam["subjects"]:
            pdf.cell(200, 10, txt=f"{sub['subject']} - {sub['obtained_marks']} / {sub['total_marks']}", ln=True)

    upload_dir = r"C:\Users\Admin\Downloads\Python Backend Final\backend\app\uploads"

    os.makedirs(upload_dir, exist_ok=True)
    pdf_path = os.path.join(upload_dir, f"{student.id}_marksheet.pdf")
    pdf.output(pdf_path)

    return FileResponse(path=pdf_path, media_type='application/pdf', filename=f"{student.first_name}_marksheet.pdf")

#------------------------------------------------------------------------------------------------------


def generate_class_progress_excel(
    class_id: int,
    section_id: int,
    exam_name: str,
    current_user: User,
    db: Session
):
    # ✅ Access control
    if current_user.user_community_type == 3:  # Class Teacher
        allocated = db.query(AllocatedMember).filter(
            AllocatedMember.user_id == current_user.id,
            AllocatedMember.allocation.has(
                class_id=class_id,
                section_id=section_id
            ),
            AllocatedMember.role == "class_teacher"
        ).first()

        if not allocated:
            return {"status":0, "detail":"Not authorized as class teacher"}

    elif current_user.user_community_type not in [1, 2]:  # Not admin or principal
        return {"status":0, "detail":"Unauthorized"}

    # ✅ Find allocation_id for class/section
    allocation = db.query(Allocation).filter(
        Allocation.class_id == class_id,
        Allocation.section_id == section_id,
        Allocation.status == 1
    ).first()

    if not allocation:
        return {"status":0, "detail":"No active allocation found"}

    # ✅ Get all students in this allocation
    students = db.query(User).join(ExamResult).filter(
        ExamResult.allocation_id == allocation.id,
        ExamResult.exam.has(exam_name=exam_name),
        User.user_community_type == 4
    ).distinct().all()

    if not students:
        return {"status":0, "detail":"Student not found"}

    # ✅ Start Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Class_{class_id}_Sec_{section_id}"

    title = "2024 - 2025 Progress Report"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    cell = ws.cell(row=1, column=1)
    cell.value = title
    cell.font = ExcelFont(size=14, bold=True)
    cell.alignment = Alignment(horizontal="center")
    ws.append([])

    headers = [
        "Roll No", "Student Name", "Exam Name", "Tamil", "English", "Maths", "Science", "Social",
        "Total", "Obtained", "Average", "Rank"
    ]
    ws.append(headers)

    subjects = ["Tamil", "English", "Mathematics", "Science", "Social Science"]
    exam_summary = defaultdict(list)

    for student in students:
        exam_results = db.query(ExamResult).join(Exam).filter(
            ExamResult.student_id == student.id,
            ExamResult.allocation_id == allocation.id,
            Exam.exam_name == exam_name
        ).all()

        for er in exam_results:
            subject_marks = db.query(
                Subject.subject_name,
                Exam.total_marks,
                Marks.obtained_marks
            ).join(Marks, Marks.subject_id == Subject.id).join(Exam).filter(
                Marks.student_id == student.id,
                Marks.exam_id == er.exam_id
            ).all()

            all_totals = db.query(ExamResult.student_id, ExamResult.total_obtained).filter(
                ExamResult.exam_id == er.exam_id,
                ExamResult.allocation_id == allocation.id
            ).order_by(desc(ExamResult.total_obtained)).all()

            rank = next((idx for idx, (sid, _) in enumerate(all_totals, start=1) if sid == student.id), None)

            subject_map = {sub: {"obtained": 0, "total": 0} for sub in subjects}
            for sub_name, total, obtained in subject_marks:
                if sub_name in subject_map:
                    subject_map[sub_name] = {"obtained": obtained, "total": total}

            obtained_total = er.total_obtained
            total_marks = er.exam.total_marks
            avg = obtained_total / len(subjects)

            exam_summary[er.exam.exam_name].append((obtained_total, total_marks))

            row = [
                student.roll_number,
                student.first_name,
                er.exam.exam_name,
                f"{subject_map['Tamil']['obtained']} / {subject_map['Tamil']['total']}",
                f"{subject_map['English']['obtained']} / {subject_map['English']['total']}",
                f"{subject_map['Mathematics']['obtained']} / {subject_map['Mathematics']['total']}",
                f"{subject_map['Science']['obtained']} / {subject_map['Science']['total']}",
                f"{subject_map['Social Science']['obtained']} / {subject_map['Social Science']['total']}",
                total_marks,
                obtained_total,
                round(avg, 2),
                rank
            ]
            ws.append(row)

    # ✅ Class average
    for exam_name, results in exam_summary.items():
        total_obt = sum(x[0] for x in results)
        total_marks = results[0][1] if results else 0
        student_count = len(results)
        avg_obt = total_obt / student_count
        avg_avg = avg_obt / len(subjects)

        avg_row = [
            "-", f"Class Average",
            exam_name,
            "-", "-", "-", "-", "-",
            total_marks,
            round(avg_obt, 2),
            round(avg_avg, 2),
            "-"
        ]
        ws.append(avg_row)

        # ✅ Auto column width
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)  # Safe way to get column letter
        max_length = 0
        for cell in col_cells:
            if not isinstance(cell, MergedCell) and cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2


    # ✅ Save and return
    upload_dir = r"C:\Users\Admin\Downloads\Python Backend Final\backend\app\uploads"
    os.makedirs(upload_dir, exist_ok=True)
    file_name = f"Progress_Report_{class_id}_{section_id}_{exam_name}.xlsx"
    file_path = os.path.join(upload_dir, file_name)
    wb.save(file_path)

    return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=file_name)


#------------------------------------CRUD for question paper-----------------------------------------------------------#

def create_bulk_question_papers(db: Session, data: list[QuestionPaperSchema]):
    now = datetime.utcnow()
    objs = [
        QuestionPaper(
            exam_id=item.exam_id,
            allocation_id=item.allocation_id,
            subject_id=item.subject_id,
            uploaded_by=item.uploaded_by,
            file_path=item.file_path,
            status=item.status,
            created_at=now,
            updated_at=now
        ) for item in data
    ]
    db.add_all(objs)
    db.commit()
    return objs

def get_all_question_papers(db: Session):
    return db.query(QuestionPaper).filter(QuestionPaper.status == 1).all()

def get_question_paper_by_id(db: Session, id: int):
    return db.query(QuestionPaper).filter(QuestionPaper.id == id).first()

def update_question_paper(db: Session, id: int, data: QuestionPaperSchema):
    obj = db.query(QuestionPaper).filter(QuestionPaper.id == id).first()
    if not obj:
        return None
    for key, value in data.dict(exclude_unset=True).items():
        setattr(obj, key, value)
    obj.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(obj)
    return obj

def soft_delete_question_paper(db: Session, id: int, updated_by: int):
    obj = db.query(QuestionPaper).filter(QuestionPaper.id == id).first()
    if obj:
        obj.status = -1
        obj.updated_at = datetime.utcnow()
        db.commit()
        return True
    return False